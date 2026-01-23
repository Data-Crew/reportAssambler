"""
Módulo para análisis masivo de laboratorios y generación de reportes consolidados.
"""
from pathlib import Path
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.lab_analyzer import LaboratoryAnalyzer
from app.report_assembler import ReportAssembler


class LaboratoryBatchAnalyzer:
    """Analiza múltiples PDFs de laboratorio y genera reportes consolidados."""
    
    def __init__(self, assembler: ReportAssembler):
        """
        Inicializa el analizador masivo.
        
        Args:
            assembler: Instancia de ReportAssembler con la configuración de rutas
        """
        self.assembler = assembler
        self.analyzer = LaboratoryAnalyzer()
    
    def find_all_laboratory_pdfs(self) -> List[Dict]:
        """
        Encuentra todos los PDFs de laboratorio disponibles.
        
        Returns:
            Lista de diccionarios con información de cada PDF encontrado:
            {
                "dni": str,
                "dni_clean": str,
                "pdf_path": Path,
                "paciente_info": Dict (si está disponible en el Excel maestro)
            }
        """
        lab_folder = self.assembler.fecha_folder / "LABORATORIO"
        if not lab_folder.exists():
            return []
        
        pdfs_encontrados = []
        
        # Buscar todos los PDFs en la carpeta LABORATORIO
        for pdf_file in lab_folder.glob("*.pdf"):
            # Extraer DNI del nombre del archivo (asumiendo formato: DNI.pdf)
            dni_clean = pdf_file.stem
            # Intentar encontrar información del paciente en el Excel maestro
            paciente_info = self._find_patient_info(dni_clean)
            
            pdfs_encontrados.append({
                "dni": paciente_info.get("DNI", dni_clean) if paciente_info else dni_clean,
                "dni_clean": dni_clean,
                "pdf_path": pdf_file,
                "paciente_info": paciente_info
            })
        
        return pdfs_encontrados
    
    def _find_patient_info(self, dni_clean: str) -> Optional[Dict]:
        """
        Busca información del paciente en el Excel maestro.
        
        Args:
            dni_clean: DNI sin puntos
            
        Returns:
            Diccionario con información del paciente o None
        """
        df_master = self.assembler.df_master
        if df_master is None or df_master.empty:
            return None
        
        # Buscar por DNI (puede tener puntos o no)
        dni_variants = [dni_clean, dni_clean.replace(".", "")]
        
        for dni_var in dni_variants:
            # Intentar buscar en diferentes formatos
            try:
                mask = (
                    df_master["DNI"].astype(str).str.replace(".", "") == dni_var
                )
                if mask.any():
                    paciente = df_master[mask].iloc[0]
                    return {
                        "DNI": paciente.get("DNI"),
                        "APELLIDOS": paciente.get("APELLIDOS", ""),
                        "NOMBRES": paciente.get("NOMBRES", ""),
                    }
            except (KeyError, AttributeError):
                continue
        
        return None
    
    def analyze_all_laboratories(self) -> List[Dict]:
        """
        Analiza todos los laboratorios disponibles.
        
        Returns:
            Lista de diccionarios con resultados del análisis de cada paciente
        """
        pdfs_encontrados = self.find_all_laboratory_pdfs()
        resultados = []
        
        for pdf_info in pdfs_encontrados:
            try:
                resultado = self.analyzer.analyze_pdf(pdf_info["pdf_path"])
                
                # Agregar información del paciente
                paciente_info = pdf_info.get("paciente_info", {})
                resultado["dni"] = pdf_info["dni"]
                resultado["dni_clean"] = pdf_info["dni_clean"]
                resultado["apellidos"] = paciente_info.get("APELLIDOS", "N/A")
                resultado["nombres"] = paciente_info.get("NOMBRES", "N/A")
                resultado["nombre_completo"] = f"{resultado['apellidos']} {resultado['nombres']}".strip()
                
                resultados.append(resultado)
            except Exception as e:
                # Si hay error, agregar información básica con error
                paciente_info = pdf_info.get("paciente_info", {})
                resultados.append({
                    "dni": pdf_info["dni"],
                    "dni_clean": pdf_info["dni_clean"],
                    "apellidos": paciente_info.get("APELLIDOS", "N/A"),
                    "nombres": paciente_info.get("NOMBRES", "N/A"),
                    "nombre_completo": f"{paciente_info.get('APELLIDOS', 'N/A')} {paciente_info.get('NOMBRES', 'N/A')}".strip(),
                    "error": str(e),
                    "total_parametros": 0,
                    "fuera_de_rango": 0,
                    "resultados": []
                })
        
        return resultados
    
    def generate_excel_report(self, resultados: List[Dict], output_path: Optional[Path] = None) -> Path:
        """
        Genera un reporte Excel consolidado con formato condicional.
        
        Args:
            resultados: Lista de resultados del análisis
            output_path: Ruta donde guardar el Excel (opcional)
            
        Returns:
            Ruta al archivo Excel generado
        """
        if output_path is None:
            fecha_str = self.assembler.fecha_folder.name
            output_path = self.assembler.fecha_folder / f"REPORTE_LABORATORIOS_{fecha_str}.xlsx"
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Laboratorios"
        
        # Colores para formato condicional
        fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_header = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        font_header = Font(bold=True, color="FFFFFF", size=11)
        font_normal = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            "DNI", "Apellidos", "Nombres", "Parámetro", "Valor", "Unidad",
            "Rango Min", "Rango Max", "Estado", "Dirección", "Exceso/Deficiencia"
        ]
        
        # Escribir encabezados
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = fill_header
            cell.font = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Datos
        row = 2
        for resultado in resultados:
            if "error" in resultado:
                # Si hay error, agregar una fila con el error
                ws.cell(row=row, column=1, value=resultado["dni"]).border = border
                ws.cell(row=row, column=2, value=resultado["apellidos"]).border = border
                ws.cell(row=row, column=3, value=resultado["nombres"]).border = border
                error_msg = resultado.get("error", "Error desconocido")
                ws.cell(row=row, column=4, value=f"ERROR: {error_msg}").border = border
                error_cell = ws.cell(row=row, column=4)
                error_cell.fill = fill_rojo
                error_cell.font = Font(bold=True, color="000000")
                # Colorear también las celdas del paciente
                ws.cell(row=row, column=1).fill = fill_rojo
                ws.cell(row=row, column=2).fill = fill_rojo
                ws.cell(row=row, column=3).fill = fill_rojo
                row += 1
                continue
            
            resultados_paciente = resultado.get("resultados", [])
            
            if not resultados_paciente:
                # Si no hay resultados, agregar fila indicando que no se encontraron parámetros
                ws.cell(row=row, column=1, value=resultado["dni"]).border = border
                ws.cell(row=row, column=2, value=resultado["apellidos"]).border = border
                ws.cell(row=row, column=3, value=resultado["nombres"]).border = border
                ws.cell(row=row, column=4, value="No se encontraron parámetros").border = border
                row += 1
                continue
            
            # Agregar cada parámetro como una fila
            for param_result in resultados_paciente:
                # DNI, Apellidos, Nombres (solo en la primera fila del paciente)
                if param_result == resultados_paciente[0]:
                    ws.cell(row=row, column=1, value=resultado["dni"]).border = border
                    ws.cell(row=row, column=2, value=resultado["apellidos"]).border = border
                    ws.cell(row=row, column=3, value=resultado["nombres"]).border = border
                else:
                    # Dejar vacío para parámetros siguientes del mismo paciente
                    ws.cell(row=row, column=1).border = border
                    ws.cell(row=row, column=2).border = border
                    ws.cell(row=row, column=3).border = border
                
                # Parámetro
                parametro_cell = ws.cell(row=row, column=4, value=param_result["parametro"])
                parametro_cell.border = border
                parametro_cell.font = font_normal
                
                # Valor
                valor = param_result.get("valor")
                if isinstance(valor, (int, float)):
                    valor_str = f"{valor:.2f}"
                else:
                    valor_str = str(valor)
                valor_cell = ws.cell(row=row, column=5, value=valor_str)
                valor_cell.border = border
                valor_cell.font = font_normal
                
                # Unidad
                unidad_cell = ws.cell(row=row, column=6, value=param_result.get("unidad", ""))
                unidad_cell.border = border
                unidad_cell.font = font_normal
                
                # Rango Min
                rango_min = param_result.get("rango_min")
                rango_min_str = f"{rango_min:.1f}" if rango_min is not None else "N/A"
                ws.cell(row=row, column=7, value=rango_min_str).border = border
                
                # Rango Max
                rango_max = param_result.get("rango_max")
                rango_max_str = f"{rango_max:.1f}" if rango_max is not None else "N/A"
                ws.cell(row=row, column=8, value=rango_max_str).border = border
                
                # Estado (con formato condicional)
                fuera_de_rango = param_result.get("fuera_de_rango", False)
                estado_cell = ws.cell(row=row, column=9, value="⚠️ Fuera" if fuera_de_rango else "✅ OK")
                estado_cell.border = border
                estado_cell.font = font_normal
                
                # Aplicar color de fondo según estado
                if fuera_de_rango:
                    estado_cell.fill = fill_amarillo
                    # También colorear la celda del valor
                    valor_cell.fill = fill_amarillo
                else:
                    estado_cell.fill = fill_verde
                    valor_cell.fill = fill_verde
                
                # Dirección
                direccion = param_result.get("direccion", "")
                direccion_str = "⬆️ Alto" if direccion == "alto" else "⬇️ Bajo" if direccion == "bajo" else ""
                ws.cell(row=row, column=10, value=direccion_str).border = border
                
                # Exceso/Deficiencia
                exceso = param_result.get("exceso")
                deficiencia = param_result.get("deficiencia")
                if exceso:
                    exc_def_str = f"{exceso:.2f}"
                elif deficiencia:
                    exc_def_str = f"{deficiencia:.2f}"
                else:
                    exc_def_str = ""
                ws.cell(row=row, column=11, value=exc_def_str).border = border
                
                row += 1
        
        # Ajustar ancho de columnas
        column_widths = {
            "A": 15,  # DNI
            "B": 25,  # Apellidos
            "C": 25,  # Nombres
            "D": 40,  # Parámetro
            "E": 12,  # Valor
            "F": 12,  # Unidad
            "G": 12,  # Rango Min
            "H": 12,  # Rango Max
            "I": 12,  # Estado
            "J": 12,  # Dirección
            "K": 15   # Exceso/Deficiencia
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Congelar primera fila
        ws.freeze_panes = "A2"
        
        # Guardar archivo
        wb.save(output_path)
        
        return output_path
